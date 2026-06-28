import scrapy
import re
import os
import asyncio
import time
from scrapy_playwright.page import PageMethod
from openpyxl import Workbook, load_workbook


XLSX_PATH = "samarinda-tomorow-rev-overnight.xlsx"
DEBUG_DIR = "debug_screenshots"


class MapsSpider(scrapy.Spider):
    name = "maps"
    _xlsx_lock = asyncio.Lock()

    async def start(self):
        os.makedirs(DEBUG_DIR, exist_ok=True)

        url = "https://www.google.com/maps/search/tomoro+coffee+samarinda"
        yield scrapy.Request(
            url,
            callback=self.parse_list,
            errback=self.handle_list_error,
            meta={
                "playwright": True,
                "playwright_page_methods": [
                    PageMethod("wait_for_selector", "a.hfpxzc"),
                    PageMethod("wait_for_timeout", 3000),
                ],
            },
        )

    def handle_list_error(self, failure):
        self.logger.error(f"[LIST] Gagal load halaman pencarian: {failure.value}")

    def parse_list(self, response):
        cards = response.css("a.hfpxzc")
        self.logger.info(f"Ketemu {len(cards)} hasil di list")

        for idx, card in enumerate(cards):
            href = card.attrib.get("href")
            if not href:
                self.logger.warning(f"[LIST] Card index {idx} gak punya href, skip.")
                continue

            yield scrapy.Request(
                href,
                callback=self.parse_detail,
                errback=self.handle_detail_error,
                meta={
                    "playwright": True,
                    "playwright_include_page": True,
                    "playwright_page_methods": [
                        PageMethod("wait_for_timeout", 2000),
                        PageMethod(
                            "evaluate",
                            """
                            () => {
                                const el = document.querySelector('h1.DUwDvf');
                                window.__nama_cafe = el ? el.innerText.trim() : null;
                            }
                            """
                        ),
                    ],
                    "card_index": idx,
                },
            )

    def handle_detail_error(self, failure):
        idx = failure.request.meta.get("card_index", "?")
        self.logger.error(
            f"[DETAIL #{idx}] Request gagal total (gak sampai parse_detail): "
            f"{failure.request.url} -> {failure.value}"
        )
        page = failure.request.meta.get("playwright_page")
        if page:
            asyncio.create_task(self._safe_close(page))

    async def _safe_close(self, page):
        try:
            await page.close()
        except Exception:
            pass

    async def _safe_screenshot(self, page, nama_cafe, step):
        try:
            safe_name = re.sub(r'[^a-zA-Z0-9_-]', "_", nama_cafe)[:50]
            path = os.path.join(DEBUG_DIR, f"{safe_name}_{step}.png")
            await page.screenshot(path=path)
            self.logger.info(f"[{nama_cafe}] Screenshot disimpan: {path}")
        except Exception as e:
            self.logger.warning(f"[{nama_cafe}] Gagal screenshot di step '{step}': {e}")

    async def _click_with_retry(self, page, locator_obj, label, nama_cafe, retries=2):
        for attempt in range(1, retries + 1):
            try:
                await locator_obj.click(timeout=2500, force=True)
                self.logger.info(f"[{nama_cafe}] Klik '{label}' sukses (percobaan {attempt})")
                return True
            except Exception as e:
                if attempt == retries:
                    self.logger.warning(f"[{nama_cafe}] Klik '{label}' gagal di percobaan terakhir: {e}")
                await page.wait_for_timeout(500)
        return False

    async def parse_detail(self, response):
        t_start = time.monotonic()
        page = response.meta["playwright_page"]
        idx = response.meta.get("card_index", "?")

        # Blokir Aset Gambar & Media agar Render Ringan
        try:
            async def intercept_route(route, request):
                resource_type = request.resource_type
                url = request.url.lower()
                if resource_type in ["image", "font", "media", "stylesheet"] or \
                   any(ext in url for ext in [".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg", ".woff", ".ttf"]):
                    await route.abort()
                elif "maps/vt" in url or "kh/v=" in url or "google-analytics" in url or "stats/client" in url:
                    await route.abort()
                else:
                    await route.continue_()

            await page.route("**/*", intercept_route)
        except Exception as e:
            self.logger.warning(f"[#{idx}] Gagal mengaktifkan route blocking: {e}")

        nama_cafe = await page.evaluate("() => window.__nama_cafe")
        nama_cafe = nama_cafe.strip() if nama_cafe else f"Unknown_Cafe_{idx}"
        
        self.logger.info(f"[{nama_cafe}] >>> MULAI parse_detail")

        collected = {}

        try:
            t0 = time.monotonic()
            tab_ulasan_locator = page.locator("button[role='tab']").filter(has_text="Ulasan").first
            
            ok = await self._click_with_retry(page, tab_ulasan_locator, "Tab Ulasan", nama_cafe)
            if not ok:
                self.logger.info(f"[{nama_cafe}] Coba fallback cari tab Ulasan via teks Inggris...")
                tab_ulasan_locator = page.locator("button[role='tab']").filter(has_text="Reviews").first
                ok = await self._click_with_retry(page, tab_ulasan_locator, "Tab Ulasan (Fallback)", nama_cafe)
                
            self.logger.info(f"[{nama_cafe}] Klik Tab Ulasan selesai dalam {time.monotonic()-t0:.1f}s")

            await page.wait_for_timeout(500)


            t0 = time.monotonic()
            tombol_urut_locator = page.locator("button").filter(has_text="Urutkan").first
            if not await tombol_urut_locator.is_visible():
                tombol_urut_locator = page.locator("button").filter(has_text="Sort").first
            if not await tombol_urut_locator.is_visible():
                tombol_urut_locator = page.locator("div.TrU0dc button").first

            ok_urutan = await self._click_with_retry(page, tombol_urut_locator, "Tombol Urutan", nama_cafe)
            self.logger.info(f"[{nama_cafe}] Klik Tombol Urutan selesai dalam {time.monotonic()-t0:.1f}s")
            
            await page.wait_for_timeout(300)

            t0 = time.monotonic()
            if ok_urutan:
                pilih_terbaru_locator = page.locator("div[role='menuitemradio']").filter(has_text="Terbaru").first
                if not await pilih_terbaru_locator.is_visible():
                    pilih_terbaru_locator = page.locator("div[role='menuitemradio']").filter(has_text="Newest").first
                    
                await self._click_with_retry(page, pilih_terbaru_locator, "Pilih Paling Terbaru", nama_cafe)
            else:
                self.logger.warning(f"[{nama_cafe}] Skip pilih 'Paling Terbaru' karena tombol Urutan gagal.")
            self.logger.info(f"[{nama_cafe}] Ganti urutan ke Terbaru selesai dalam {time.monotonic()-t0:.1f}s")

            await page.wait_for_timeout(800)

            t0 = time.monotonic()
            collected = await self._extract_loop(page, nama_cafe)
            self.logger.info(f"[{nama_cafe}] _extract_loop selesai dalam {time.monotonic()-t0:.1f}s, total {len(collected)} review")

        except Exception as e:
            self.logger.error(f"[{nama_cafe}] EXCEPTION tak terduga saat scraping: {e}", exc_info=True)
            await self._safe_screenshot(page, nama_cafe, "unexpected_exception")

        finally:
            try:
                await page.close()
            except Exception as e:
                self.logger.warning(f"[{nama_cafe}] Gagal close page: {e}")

        rows = list(collected.values())

        async with self._xlsx_lock:
            self.save_to_xlsx(nama_cafe, rows)

        self.logger.info(f"[{nama_cafe}] <<< TOTAL parse_detail: {time.monotonic()-t_start:.1f}s")

        for row in rows:
            item = {"nama_cafe": nama_cafe}
            item.update(row)
            yield item

    @staticmethod
    def _normalize_key(raw_key):
        if not raw_key:
            return None
        key = re.sub(r"\s+", " ", raw_key).strip().lower()
        if not key:
            return None
        return key[0].upper() + key[1:]

    async def _get_scroll_handle(self, page, nama_cafe):
        try:
            handle = await page.evaluate_handle("""
                () => {
                    const els = document.querySelectorAll(".m6QErb");
                    let best = null, bestRatio = 0;
                    for (const el of els) {
                        if (el.scrollHeight > el.clientHeight + 50 && el.clientHeight > 100) {
                            const label = (el.getAttribute("aria-label") || "").toLowerCase();
                            if (label.includes("hasil untuk") || label.includes("results for")) continue;
                            if (el.closest('[aria-label*="Hasil untuk"], [aria-label*="results for"]')) continue;
                            const rect = el.getBoundingClientRect();
                            if (rect.x < window.innerWidth * 0.45 && rect.width > 200) {
                                const ratio = el.scrollHeight / el.clientHeight;
                                if (ratio > bestRatio) { bestRatio = ratio; best = el; }
                            }
                        }
                    }
                    return best;
                }
            """)
            el = handle.as_element()
            if el is None:
                self.logger.warning(f"[{nama_cafe}] Gagal dapat handle container scroll.")
                return None
            return handle
        except Exception as e:
            self.logger.warning(f"[{nama_cafe}] Error ambil scroll handle: {e}")
            return None

    async def _scroll_via_handle(self, handle, nama_cafe, delta=850):
        try:
            result = await handle.evaluate("""
                (el, d) => {
                    if (!el.isConnected) return "detached";
                    const before = el.scrollTop;
                    el.scrollTop += d;
                    el.dispatchEvent(new Event('scroll', { bubbles: true }));
                    return el.scrollTop !== before ? "moved" : "stuck";
                }
            """, delta)
            return result
        except Exception as e:
            self.logger.warning(f"[{nama_cafe}] Error evaluate scroll handle: {e}")
            return "error"

    async def _extract_loop(self, page, nama_cafe):
        collected = {}
        no_new_count = 0

        scroll_handle = await self._get_scroll_handle(page, nama_cafe)
        if scroll_handle is None:
            self.logger.error(f"[{nama_cafe}] Gak nemu container scroll sama sekali, batal extract loop.")
            return collected

        async def _do_scroll(iteration_idx, delta_y=850):
            nonlocal scroll_handle
            status = await self._scroll_via_handle(scroll_handle, nama_cafe, delta_y)
            if status in ("detached", "error"):
                self.logger.warning(f"[{nama_cafe}] Handle scroll {status}, re-fetch handle...")
                scroll_handle = await self._get_scroll_handle(page, nama_cafe)
                if scroll_handle is None:
                    return False
                status = await self._scroll_via_handle(scroll_handle, nama_cafe, delta_y)
            return status == "moved"
        
        self.logger.info(f"[{nama_cafe}] Warm-up scroll...")
        for attempt in range(3):
            moved = await _do_scroll(attempt, delta_y=600)
            await page.wait_for_timeout(250)
            if not moved and attempt >= 1:
                break
        await page.wait_for_timeout(800)

        click_expand_buttons_js = """
        () => {
            const buttons = document.querySelectorAll('button[aria-label="Lihat lainnya"], button.w8nwRe.kyuRq');
            let clickedCount = 0;
            buttons.forEach(btn => {
                if (btn && btn.innerText.trim().includes("Lainnya")) {
                    btn.click();
                    clickedCount++;
                }
            });
            return clickedCount;
        }
        """

        extract_js = r"""
        () => {
            const out = [];
            document.querySelectorAll("div[data-review-id]").forEach(el => {
                const id = el.getAttribute("data-review-id");
                const namaEl = el.querySelector("div.d4r55");
                const reviewEl = el.querySelector("span.wiI7pd");
                const ratingEl = el.querySelector("span.kvMYJc");
                const waktuEl = el.querySelector("span.rsqaWe");

                let rating = null;
                if (ratingEl) {
                    const label = ratingEl.getAttribute("aria-label") || "";
                    const match = label.match(/\d+/);
                    rating = match ? match[0] : null;
                }

                const extra = {};
                el.querySelectorAll("div.PBK6be").forEach(pbk => {
                    const boldEl = pbk.querySelector("span[style*='font-weight:bold'], span[style*='font-weight: bold'], b");
                    if (boldEl) {
                        let key = boldEl.innerText.trim().replace(/:$/, "");
                        let value = null;
                        const labelContainer = boldEl.closest("div") || boldEl.parentElement;
                        const containerFullText = labelContainer ? labelContainer.innerText.trim() : "";
                        let remainder = containerFullText.replace(boldEl.innerText.trim(), "").trim();
                        remainder = remainder.replace(/^:\s*/, "");

                        if (remainder) {
                            value = remainder;
                        } else {
                            const directDivs = Array.from(pbk.querySelectorAll(":scope > div"));
                            const labelDivIndex = directDivs.findIndex(d => d.contains(boldEl));
                            if (labelDivIndex !== -1 && directDivs[labelDivIndex + 1]) {
                                value = directDivs[labelDivIndex + 1].innerText.trim();
                            }
                        }
                        if (key) extra[key] = value || null;
                    }
                });

                out.push({
                    id: id,
                    nama_reviewer: namaEl ? namaEl.innerText.trim() : null,
                    review: reviewEl ? reviewEl.innerText.trim() : null,
                    rating: rating,
                    waktu_relatif: waktuEl ? waktuEl.innerText.trim() : null,
                    extra: extra
                });
            });
            return out;
        }
        """

        self.logger.info(f"[{nama_cafe}] Memulai Infinite Turbo Loop (Max 1jt Tarikan Scroll)...")
        
        for i in range(1000000):
            try:
                expanded = await page.evaluate(click_expand_buttons_js)
                if expanded > 0:
                    await page.wait_for_timeout(100)
            except Exception:
                pass

            try:
                results = await page.evaluate(extract_js)
            except Exception as e:
                self.logger.error(f"[{nama_cafe}] Gagal evaluate extract_js di iterasi {i}: {e}")
                break

            new_count = 0
            for r in results:
                rid = r["id"]
                if rid and (r["nama_reviewer"] or r["review"]):
                    if rid not in collected:
                        new_count += 1
                        row = {
                            "nama_reviewer": r["nama_reviewer"],
                            "rating": r["rating"],
                            "waktu_relatif": r["waktu_relatif"],
                            "review": r["review"],
                        }
                        for raw_key, val in r["extra"].items():
                            norm_key = self._normalize_key(raw_key)
                            if norm_key:
                                row[norm_key] = val
                        collected[rid] = row

            self.logger.info(f"[{nama_cafe}] Iterasi {i}: total {len(collected)} (+{new_count} baru)")


            no_new_count = 0 if new_count else no_new_count + 1
            if no_new_count >= 14:
                self.logger.info(f"[{nama_cafe}] Stop: Ulasan sudah habis dikuras (14x scroll kosong).")
                break

            try:
                scrolled = await _do_scroll(i, delta_y=850)
            except Exception:
                break

            await page.wait_for_timeout(1200)

        return collected

    def save_to_xlsx(self, nama_cafe, rows):
        if not rows:
            self.logger.warning(f"Gak ada data buat {nama_cafe}, skip simpan sheet.")
            return

        safe_name = re.sub(r'[\\/?*\[\]:]', "_", nama_cafe).strip()
        sheet_name = safe_name[:31] if safe_name else "Unknown"

        if os.path.exists(XLSX_PATH):
            wb = load_workbook(XLSX_PATH)
        else:
            wb = Workbook()
            default_sheet = wb["Sheet"]
            wb.remove(default_sheet)

        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])

        ws = wb.create_sheet(title=sheet_name)

        fieldnames = []
        for row in rows:
            for key in row.keys():
                if key not in fieldnames:
                    fieldnames.append(key)

        ws.append(fieldnames)
        for row in rows:
            ws.append([row.get(field, "") for field in fieldnames])

        wb.save(XLSX_PATH)
        self.logger.info(f"Tersimpan ke sheet '{sheet_name}' di {XLSX_PATH} ({len(rows)} baris)")